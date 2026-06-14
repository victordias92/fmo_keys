import logging
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from database.connection import data_dir

logger = logging.getLogger(__name__)

FAILURES_EXCEL_FILENAME = "falhas_exportadas.xlsx"

HEADERS = [
    "Local",
    "Descricao",
    "Responsavel",
    "Data",
    "Hora",
    "Dia",
    "Status",
    "Observacoes",
    "Possui imagem",
]

fillgrey = PatternFill("solid", "DDDDDD")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _excel_path() -> Path:
    return data_dir / FAILURES_EXCEL_FILENAME


def format_cell(cell, value, fill=None, font=None):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if font:
        cell.font = font


def _setup_headers(ws) -> None:
    ws.append(HEADERS)
    for col_letter, width in zip("ABCDEFGHI", [30, 40, 25, 15, 12, 15, 15, 30, 15]):
        ws.column_dimensions[col_letter].width = width

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C0504D")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _create_workbook() -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Falhas"
    _setup_headers(ws)
    return wb, ws


def _load_or_create_workbook(excel_file: Path) -> tuple[Workbook, object]:
    if excel_file.exists():
        try:
            wb = load_workbook(excel_file)
            return wb, wb.active
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            logger.warning("Arquivo Excel de falhas invalido (%s). Recriando.", exc)
            excel_file.unlink(missing_ok=True)
    return _create_workbook()


def ensure_export_workbook() -> Path:
    excel_file = _excel_path()
    wb, ws = _load_or_create_workbook(excel_file)
    if ws.max_row >= 1:
        current_headers = [ws.cell(1, col).value for col in range(1, len(HEADERS) + 1)]
        if current_headers != HEADERS:
            wb, ws = _create_workbook()
    wb.save(excel_file)
    return excel_file


def append_failure(
    location: str,
    description: str,
    reported_by: str,
    failure_date: str,
    failure_time: str,
    failure_day: str,
    status: str,
    notes: str,
    has_image: bool,
) -> None:
    excel_file = _excel_path()
    wb, ws = _load_or_create_workbook(excel_file)
    row_data = [
        location,
        description,
        reported_by,
        failure_date,
        failure_time,
        failure_day,
        status,
        notes,
        "sim" if has_image else "nao",
    ]
    ws.append(row_data)
    for col in range(1, len(row_data) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        format_cell(cell, row_data[col - 1])
        cell.font = Font(size=12)
        if ws.max_row % 2 == 0:
            cell.fill = fillgrey
    wb.save(excel_file)
