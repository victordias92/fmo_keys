import logging
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from database.connection import data_dir

logger = logging.getLogger(__name__)

EXCEL_FILENAME = "chaves_exportadas.xlsx"

HEADERS = [
    "Chave",
    "Ultimo usuario",
    "Data",
    "Hora",
    "Dia",
    "Status",
    "Data",
    "Hora",
    "dia",
]

fillgrey = PatternFill("solid", "DDDDDD")
fillgreen = PatternFill("solid", fgColor="00FF00")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
font_devolvida = Font(bold=True, color="FFFFFF")


def _excel_path() -> Path:
    return data_dir / EXCEL_FILENAME


def format_cell(cell, value, fill=None, font=None):
    """Aplica valor e estilo padrao em uma celula."""
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if font:
        cell.font = font


def _setup_headers(ws) -> None:
    ws.append(HEADERS)
    for col_letter, width in zip("ABCDEFGHI", [40, 25, 25, 25, 25, 25, 25, 25, 25]):
        ws.column_dimensions[col_letter].width = width

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _create_workbook() -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    _setup_headers(ws)
    return wb, ws


def _load_or_create_workbook(excel_file: Path) -> tuple[Workbook, object]:
    if excel_file.exists():
        try:
            wb = load_workbook(excel_file)
            return wb, wb.active
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            logger.warning(
                "Arquivo Excel invalido (%s). Recriando: %s",
                exc,
                excel_file,
            )
            excel_file.unlink(missing_ok=True)

    return _create_workbook()


def ensure_export_workbook() -> Path:
    """Garante que o historico formatado existe e retorna o caminho para download."""
    excel_file = _excel_path()
    wb, ws = _load_or_create_workbook(excel_file)
    if ws.max_row >= 1:
        current_headers = [ws.cell(1, col).value for col in range(1, len(HEADERS) + 1)]
        if current_headers != HEADERS:
            logger.warning(
                "Planilha com cabecalho incompativel (%s). Recriando historico formatado.",
                current_headers,
            )
            wb, ws = _create_workbook()
    wb.save(excel_file)
    return excel_file


def marcar_devolucao(ws, row, used_date, used_time, used_day):
    format_cell(ws.cell(row=row, column=6), "devolvida", fill=fillgreen, font=font_devolvida)
    format_cell(ws.cell(row=row, column=7), used_date)
    format_cell(ws.cell(row=row, column=8), used_time)
    format_cell(ws.cell(row=row, column=9), used_day)


def get_last_user_key(
    keys: str,
    used_date: str,
    used_time: str,
    used_day: str,
):
    excel_file = _excel_path()
    if not excel_file.exists():
        return

    wb, ws = _load_or_create_workbook(excel_file)

    for row in range(2, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=1).value or "")
        chave = "".join(c for c in cell_value if c.isdigit())

        if chave == str(keys):
            marcar_devolucao(ws, row, used_date, used_time, used_day)

    wb.save(excel_file)


def append_usage(
    key_name: str,
    user: str,
    used_date: str,
    used_time: str,
    used_day: str,
    on_use: str,
):
    excel_file = _excel_path()
    wb, ws = _load_or_create_workbook(excel_file)

    dados = [key_name, user, used_date, used_time, used_day]
    ws.append([key_name, user, used_date, used_time, used_day, on_use])

    for col in range(1, len(dados) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        format_cell(cell, dados[col - 1])
        cell.font = Font(size=12)
        if ws.max_row % 2 == 0:
            cell.fill = fillgrey

    wb.save(excel_file)
